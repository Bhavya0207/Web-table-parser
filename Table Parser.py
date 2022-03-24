import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import openpyxl
from openpyxl import Workbook
wb = openpyxl.Workbook() 
sheet = wb.active 
page = requests.get("https://www.bankexamstoday.com/2017/03/states-their-chief-minister-governor.html")
soup=bs(page.content,'html.parser')
#print(soup.prettify())
h=[]
h2=[]
headings=soup.find_all('th')
headings2=headings[-3:]
headings=headings[:-3]
for i in headings:
    h.append(i.get_text())
for i in headings2:
    h2.append(i.get_text())
text=soup.find_all('tr')
text2=text[-8:]
text=text[:-8]
d={}
li=[]
lix=[]
text.pop(0)
for i in text:
    li2=[]
    data=i.find_all('td')
    #print(data)
    
    for x in range(len(data)):
        li2.append(data[x].get_text())
    li.append(li2)
for i in text2:
    li3=[]
    data=i.find_all('td')
    #print(data)
    
    for x in range(len(data)):
        li3.append(data[x].get_text())
    lix.append(li3)
#print(lix)
for i in range(len(h)):
    h1=sheet.cell(row=1,column=i+1)
    h1.value=h[i]
for i in range(len(h2)):
    h1=sheet.cell(row=30,column=i+1)
    h1.value=h[i]
for i in range(len(li)):
    for x in range(len(h)):
        if i==0:
            z=2
        else:
            z=i+1
        cell=sheet.cell(row=z,column=x+1)
        try:
            cell.value=li[i][x]
        except IndexError:
            pass
        #print(cell.value)

for i in range(len(lix)):
    for x in range(len(h2)):
        if i==0:
            z=31
        else:
            z=31+i
        cell=sheet.cell(row=z,column=x+1)
        try:
            cell.value=lix[i][x]
        except IndexError:
            pass
        #print(cell.value)

wb.save("h.xlsx")